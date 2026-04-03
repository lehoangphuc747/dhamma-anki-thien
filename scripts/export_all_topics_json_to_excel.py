# -*- coding: utf-8 -*-
"""
Xuất Excel từ `Dhamma_Anki_by_topic_json/all_topics.json`.

Chạy từ thư mục Thiền:
    python scripts/export_all_topics_json_to_excel.py
"""

from __future__ import annotations

import json
from pathlib import Path

import pandas as pd


BASE_DIR = Path(__file__).resolve().parent.parent
INPUT_JSON = BASE_DIR / "Dhamma_Anki_by_topic_json" / "all_topics.json"
OUTPUT_XLSX = BASE_DIR / "Dhamma_Anki_by_topic_json" / "all_topics_export.xlsx"


def _autofit_and_style_xlsx(xlsx_path: Path, df_len: int) -> None:
    """Format Excel cho dễ đọc: header đậm, wrap text, autofit width, freeze panes."""
    from openpyxl import load_workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    wb = load_workbook(xlsx_path)
    ws = wb.active

    # Tên cột ở hàng 1
    col_names = [cell.value for cell in ws[1]]
    last_col = ws.max_column
    last_row = 1 + max(df_len, 0)

    header_fill = PatternFill(start_color="FFEFEFEF", end_color="FFEFEFEF", fill_type="solid")
    header_font = Font(bold=True, name="Calibri")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Style header + căn giữa
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_align

    # Cố định dòng tiêu đề
    ws.freeze_panes = "A2"

    # Bật filter
    ws.auto_filter.ref = f"A1:{ws.cell(row=last_row, column=last_col).coordinate}"

    # Wrap text cho các cột văn bản dài
    wrap_align = Alignment(horizontal="left", vertical="top", wrap_text=True)
    target_wrap_cols = {"vietnamese", "explain", "title"}

    # Chỉ áp dụng wrap cho các cột thuộc nhóm này
    for row in ws.iter_rows(min_row=2, max_row=last_row, min_col=1, max_col=last_col):
        for cell in row:
            header = col_names[cell.column - 1] if cell.column - 1 < len(col_names) else None
            if header in target_wrap_cols:
                cell.alignment = wrap_align

    # Tự canh độ rộng cột (giới hạn để không làm xấu layout)
    max_width_cap = 80
    for col in range(1, last_col + 1):
        max_len = 0
        for r in range(1, last_row + 1):
            val = ws.cell(row=r, column=col).value
            if val is None:
                continue
            max_len = max(max_len, len(str(val)))
        width = min(max_len + 2, max_width_cap)
        ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = width

    wb.save(xlsx_path)


def main() -> None:
    if not INPUT_JSON.exists():
        raise FileNotFoundError(f"Không tìm thấy file: {INPUT_JSON}")

    raw = INPUT_JSON.read_text(encoding="utf-8")
    data = json.loads(raw)
    if not isinstance(data, list):
        raise ValueError("all_topics.json không phải danh sách (list).")

    # Chuẩn hóa schema theo format note bạn dùng.
    # JSON hiện có: id, english, ipa, vietnamese, explain, title
    rows: list[dict[str, str]] = []
    for item in data:
        if not isinstance(item, dict):
            continue
        rows.append(
            {
                "id": str(item.get("id", "")),
                "english": str(item.get("english", "")),
                "ipa": str(item.get("ipa", "")),
                "vietnamese": str(item.get("vietnamese", "")),
                "explain": str(item.get("explain", "")),
                "title": str(item.get("title", "")),
                # Audio trong JSON hiện chưa có; để cột trống cho đồng bộ.
                "audio_word": str(item.get("audio_word", "")) if "audio_word" in item else "",
                "audio_vietnamese": str(item.get("audio_vietnamese", ""))
                if "audio_vietnamese" in item
                else "",
            }
        )

    df = pd.DataFrame(rows, columns=[
        "id",
        "english",
        "ipa",
        "vietnamese",
        "explain",
        "title",
        "audio_word",
        "audio_vietnamese",
    ])

    # Ghi Excel UTF-8/Unicode tốt.
    df.to_excel(OUTPUT_XLSX, index=False)
    # Windows console đôi khi dùng encoding cp1252 nên in đường dẫn có ký tự Unicode có thể lỗi.
    print("Wrote:", OUTPUT_XLSX.name)
    print("Rows:", len(df))

    # Thêm format để dễ đọc
    _autofit_and_style_xlsx(OUTPUT_XLSX, len(df))


if __name__ == "__main__":
    main()

